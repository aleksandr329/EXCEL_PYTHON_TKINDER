import openpyxl
from tkinter import *
from tkinter.ttk import Radiobutton
from constants import *
from function import *
from openpyxl import load_workbook



def clicked():
    if selected.get() == 1:
    	scan()

    if selected.get() == 2:
        def zapis():  #<-  Функция для записи данных
            file = 'table/Таблица учета Оплат.xlsx'
            wb = load_workbook(file)
            ws = wb['Лист1']
            number = scan2()
            ws['A' + str(number)] = number - 1
            ws['B' + str(number)] = txt.get()
            ws['C' + str(number)] = txt2.get()
            ws['D' + str(number)] = txt3.get()
            ws['E' + str(number)] = int(txt4.get())
            ws['F' + str(number)] = txt5.get()
            ws['G' + str(number)] = txt6.get()
            ws['H' + str(number)] = txt7.get()
            ws['I' + str(number)] = txt8.get()
            wb.save(file)
            wb.close()

        txt = Entry(root,width=20)
        btn = Button(root, text="ОК", command=zapis)
        lbl = Label(root, text='Введите дату счета')
        lbl.grid(column=0, row=1)
        txt.grid(column=1, row=1)
        btn.grid(column=2, row=5)
        txt2 = Entry(root,width=20)
        lbl2 = Label(root, text='Введите номер счета')
        txt2.grid(column=1, row=2)
        lbl2.grid(column=0, row=2)
        txt3 = Entry(root,width=20)
        lbl3 = Label(root, text='Введите имя организации')
        txt3.grid(column=1, row=3)
        lbl3.grid(column=0, row=3)
        txt4 = Entry(root,width=20)
        lbl4 = Label(root, text='Введите инн организации')
        txt4.grid(column=1, row=4)
        lbl4.grid(column=0, row=4)
        txt5 = Entry(root,width=20)
        lbl5 = Label(root, text='Введите всю сумму платежа')
        txt5.grid(column=1, row=5)
        lbl5.grid(column=0, row=5)
        txt6 = Entry(root,width=20)
        lbl6 = Label(root, text='Введите остаток сколько нужно доплатить')
        txt6.grid(column=1, row=6)
        lbl6.grid(column=0, row=6)
        txt7 = Entry(root,width=20)
        lbl7 = Label(root, text='Введите дату до которой нужно внести остаток суммы')
        txt7.grid(column=1, row=7)
        lbl7.grid(column=0, row=7)
        txt8 = Entry(root,width=20)
        lbl8 = Label(root, text='Введите статус Оплачен или Не оплачен')
        txt8.grid(column=1, row=8)
        lbl8.grid(column=0, row=8)

    if selected.get() == 3:
        def name():
            scan3(txt.get())

        txt = Entry(root,width=20)
        btn = Button(root, text="ОК", command=name)
        lbl = Label(root, text='Введите название  ОРГАНИЗАЦИИ')
        lbl.grid(column=0, row=4)
        txt.grid(column=1, row=4)
        btn.grid(column=2, row=4)

    if selected.get() == 4:
        def name():
            scan4(int(txt.get()))

        txt = Entry(root,width=20)
        btn = Button(root, text="ОК", command=name)
        lbl = Label(root, text='Введите ИНН')
        lbl.grid(column=0, row=4)
        txt.grid(column=1, row=4)
        btn.grid(column=2, row=4)

    if selected.get() == 5:

        def name3():

            def name2():

                def name_stutus():
                    if txt3.get().lower() in 'оплачен' or txt3.get().lower() in 'не оплачен':
                        status(txt.get(), txt3.get())
                    else:
                        name2()

                def name_ostatok():
                    ostatok(txt.get(), txt3.get())

                def name_data():
                    new_date(txt.get(), txt3.get())

                if txt2.get().lower() in 'статус':
                    txt3 = Entry(root,width=20)
                    btn3 = Button(root, text="ОК", command=name_stutus)
                    lbl3 = Label(root, text='Какой статус поставить?')
                    lbl3.grid(column=0, row=6)
                    txt3.grid(column=1, row=6)
                    btn3.grid(column=2, row=6)


                if txt2.get().lower() in 'остаток':
                    txt3 = Entry(root,width=20)
                    btn3 = Button(root, text="ОК", command=name_ostatok)
                    lbl3 = Label(root, text='Какой остаток поставить?')
                    lbl3.grid(column=0, row=6)
                    txt3.grid(column=1, row=6)
                    btn3.grid(column=2, row=6)

                if txt2.get().lower() in 'дату':
                    txt3 = Entry(root,width=20)
                    btn3 = Button(root, text="ОК", command=name_data)
                    lbl3 = Label(root, text='Какую дату поставить?')
                    lbl3.grid(column=0, row=6)
                    txt3.grid(column=1, row=6)
                    btn3.grid(column=2, row=6)

                else:
                    name3()

            scan5(txt.get())
            txt2 = Entry(root,width=20)
            btn2 = Button(root, text="ОК", command=name2)
            lbl2 = Label(root, text=vopros1)
            lbl2.grid(column=0, row=5)
            txt2.grid(column=1, row=5)
            btn2.grid(column=2, row=5)

        txt = Entry(root,width=20)
        btn = Button(root, text="ОК", command=name3)
        lbl = Label(root, text='Введите номер СЧЕТА')
        lbl.grid(column=0, row=4)
        txt.grid(column=1, row=4)
        btn.grid(column=2, row=4)
        
       
root = Tk()
root.title("Приложение СРЕДСТВА ЗАЩИТЫ")
root.geometry('700x400')
root.resizable(height = False, width = False)
root.iconphoto(True, PhotoImage(file=('images/iconka.png')))
selected = IntVar()

rad1 = Radiobutton(root,text='ЗАДОЛЖНОСТЬ', value=1, variable=selected)
rad2 = Radiobutton(root,text='ЗАПИСАТЬ', value=2, variable=selected)
rad3 = Radiobutton(root,text='ОРГАНИЗАЦИЯ', value=3, variable=selected)
rad4 = Radiobutton(root,text='ИНН', value=4, variable=selected)
rad5 = Radiobutton(root,text='СЧЕТ', value=5, variable=selected)
btn = Button(root, text="ОК", command=clicked)
lbl = Label(root)
rad1.grid(column=0, row=0)
rad2.grid(column=1, row=0)
rad3.grid(column=2, row=0)
rad4.grid(column=3, row=0)
rad5.grid(column=4, row=0)
btn.grid(column=5, row=0)
lbl.grid(column=0, row=1)

root.mainloop()
