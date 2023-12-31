import openpyxl
from openpyxl import load_workbook
from constants import time_now, vopros1
from pdf import pdf_file



def scan():  #<- Функция для сканирования и вывода информации в консоль и в файл
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'I{i}'].value is None:
            break

        if sheet[f'I{i}'].value in 'Не оплачен':
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            info = f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo} \n'''
            print(info)
            with open(f'C:\\Users\\User\\Desktop\\Отчет {time_now}.txt', 'a') as file_txt:
                file_txt.write(info)
    pdf_file() #<-- Функция для записи в pdf формате отчета

def scan2():  #<- Функция для определения пустой клетки для новой записи
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'A{i}'].value is None:
            return i
            break


def scan3(name):  #<- Функция для скана по имени фирмы
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'D{i}'].value is None:
            break

        if sheet[f'D{i}'].value.lower() in name.lower():
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            info = f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo} \n'''
            print(info)
            with open(f'C:\\Users\\User\\Desktop\\Отчет по названию фирмы {time_now}.txt', 'a') as file_txt:
                file_txt.write(info)
            continue

        else:
            name_tabl = sheet[f'D{i}'].value.lower()
            name_user = name.lower()
            try:
                if name_tabl[0] in name_user[0] and name_tabl[1] in name_user[1] and name_tabl[2] in name_user[2]:
                    i_d = sheet[f'A{i}'].value
                    number = sheet[f'C{i}'].value
                    organizace = sheet[f'D{i}'].value
                    inn = sheet[f'E{i}'].value
                    summa = sheet[f'F{i}'].value
                    ostatok = sheet[f'G{i}'].value
                    chislo = sheet[f'H{i}'].value
                    info = f'''ID {i_d}, Номер счета: {number},
                    ИНН: {inn}, Организация: {organizace},
                    Общая сумма: {summa}, Остаток: {ostatok},
                    Оплатить до: {chislo} \n'''
                    print(info)
                    with open(f'C:\\Users\\User\\Desktop\\Отчет по названию фирмы {time_now}.txt', 'a') as file_txt:
                        file_txt.write(info)
                    continue
            
                if name_tabl[0] in name_user[0] and name_tabl[1] in name_user[1] and name_tabl[2] in name_user[2] and name_tabl[3] in name_user[3]:
                    i_d = sheet[f'A{i}'].value
                    number = sheet[f'C{i}'].value
                    organizace = sheet[f'D{i}'].value
                    inn = sheet[f'E{i}'].value
                    summa = sheet[f'F{i}'].value
                    ostatok = sheet[f'G{i}'].value
                    chislo = sheet[f'H{i}'].value
                    info = f'''ID {i_d}, Номер счета: {number},
                    ИНН: {inn}, Организация: {organizace},
                    Общая сумма: {summa}, Остаток: {ostatok},
                    Оплатить до: {chislo} \n'''
                    print(info)
                    with open(f'C:\\Users\\User\\Desktop\\Отчет по названию фирмы {time_now}.txt', 'a') as file_txt:
                        file_txt.write(info)
                    continue

            except IndexError:
                print('Вы ввели недостаточно символов чтобы определить название фирмы')
                break

        
def scan4(inn):  #<- Функция для скана по инн фирмы
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'E{i}'].value is None:
            break

        if sheet[f'E{i}'].value == inn:
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            status = sheet[f'I{i}'].value
            info = f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo}, Статус: {status} \n'''
            print(info)
            with open(f'C:\\Users\\User\\Desktop\\Отчет по ИНН фирмы {time_now}.txt', 'a') as file_txt:
                file_txt.write(info)

def scan5(number):  #<- Функция для скана по номеру счета
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'C{i}'].value is None:
            break

        if sheet[f'C{i}'].value.lower() in number.lower():
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            info = (f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo}''')
            with open(f'C:\\Users\\User\\Desktop\\Отчет по ИНН фирмы {time_now}.txt', 'a') as file_txt:
                file_txt.write(info)

def status(number, status):
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'C{i}'].value is None:
            break

        if sheet[f'C{i}'].value.lower() in number.lower():
            wb = load_workbook(file)
            ws = wb['Лист1']
            ws['I' + str(i)] = status.capitalize()
            wb.save(file)
            wb.close()

def ostatok(number, ostatok):
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'C{i}'].value is None:
            break
                    
        if sheet[f'C{i}'].value.lower() in number.lower():
            wb = load_workbook(file)
            ws = wb['Лист1']
            ws['G' + str(i)] = ostatok
            wb.save(file)
            wb.close()

def new_date(number, data_new):
    file = 'table/Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'C{i}'].value is None:
            break
        if sheet[f'C{i}'].value.lower() in number.lower():
            wb = load_workbook(file)
            ws = wb['Лист1']
            ws['H' + str(i)] = data_new
            wb.save(file)
            wb.close()
